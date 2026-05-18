"""
Utilitário para exportação de dados em formato Excel (.xlsx).
Utiliza openpyxl.
"""
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def exportar_excel(nome_arquivo, colunas, linhas, titulo_planilha='Dados'):
    """
    Gera um HttpResponse com conteúdo Excel (.xlsx) pronto para download.

    Args:
        nome_arquivo (str): Nome do arquivo sem extensão. Ex: 'ordens_servico'
        colunas (list[str]): Lista de cabeçalhos.
        linhas (list[list]): Lista de listas com os valores de cada linha.
        titulo_planilha (str): Nome da aba da planilha.

    Returns:
        HttpResponse com Content-Type xlsx e Content-Disposition attachment.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = titulo_planilha

    # ── Estilos ──────────────────────────────────────────────────────
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0'),
    )
    data_font = Font(name='Calibri', size=10)
    alt_fill = PatternFill(start_color='F2F6FA', end_color='F2F6FA', fill_type='solid')

    # ── Cabeçalho ────────────────────────────────────────────────────
    for col_idx, titulo in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── Dados ────────────────────────────────────────────────────────
    for row_idx, linha in enumerate(linhas, start=2):
        for col_idx, valor in enumerate(linha, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            # Zebra striping
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # ── Auto-ajuste de largura das colunas ───────────────────────────
    for col_idx, titulo in enumerate(colunas, start=1):
        max_len = len(str(titulo))
        for row_idx in range(2, len(linhas) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        # Limita a 50 caracteres para não explodir a largura
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    # ── Congela painel do cabeçalho ──────────────────────────────────
    ws.freeze_panes = 'A2'

    # ── Gera resposta HTTP ───────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}.xlsx"'
    return response
